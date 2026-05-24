from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session

from app.auth.security import get_current_user
from app.core.database import get_db
from app.models.equipment import ActivityLog
from app.models.user import User

router = APIRouter(prefix="/api", tags=["Common"])


@router.get("/users/lookup")
def user_lookup(db: Session = Depends(get_db), _token: dict = Depends(get_current_user)):
    users = db.query(User).filter(User.is_active == True).order_by(User.full_name).all()
    return [{"id": u.id, "username": u.username, "full_name": u.full_name, "role": u.role} for u in users]


@router.get("/activity-logs")
def activity_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    module: str | None = Query(None),
    db: Session = Depends(get_db),
    _token: dict = Depends(get_current_user),
):
    query = db.query(ActivityLog)
    if module:
        query = query.filter(ActivityLog.module == module)
    total = query.count()
    logs = query.order_by(ActivityLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total,
        "items": [
            {
                "id": l.id,
                "user_id": l.user_id,
                "username": l.user.username if l.user else l.username,
                "action": l.action,
                "module": l.module,
                "description": l.description,
                "created_at": str(l.created_at),
            }
            for l in logs
        ],
    }


@router.get("/overdue")
def get_overdue(db: Session = Depends(get_db), _token: dict = Depends(get_current_user)):
    from datetime import datetime
    from app.models.ncr import NCR, CAPA

    now = datetime.now()
    open_ncrs = db.query(NCR).filter(NCR.status.in_(["open", "investigating"])).count()
    resolved_ncrs = db.query(NCR).filter(NCR.status == "resolved").count()
    closed_ncrs = db.query(NCR).filter(NCR.status == "closed").count()

    overdue_ncrs = db.query(NCR).filter(
        NCR.status.in_(["open", "investigating"]),
        NCR.due_date.isnot(None),
        NCR.due_date < now,
    ).count()

    overdue_capas = db.query(CAPA).filter(
        CAPA.status.in_(["open", "in_progress"]),
        CAPA.due_date.isnot(None),
        CAPA.due_date < now,
    ).count()

    return {
        "open_ncrs": open_ncrs, "resolved_ncrs": resolved_ncrs, "closed_ncrs": closed_ncrs,
        "overdue_ncrs": overdue_ncrs, "overdue_capas": overdue_capas,
    }


@router.get("/aql-table")
def aql_table():
    return {
        "levels": [
            {"label": "G-I (Giam nhe)", "value": "GI"},
            {"label": "G-II (Thuong)", "value": "GII"},
            {"label": "G-III (Chat che)", "value": "GIII"},
            {"label": "S-1 (Dac biet 1)", "value": "S1"},
            {"label": "S-2 (Dac biet 2)", "value": "S2"},
            {"label": "S-3 (Dac biet 3)", "value": "S3"},
            {"label": "S-4 (Dac biet 4)", "value": "S4"},
        ],
        "sample_size_codes": {
            "S1": [None, "A", "A", "A", "A", "A", "A", "A", "A", "A", "A", "A", "A", "B", "B", "C", "D", "E", "E"],
            "S2": [None, "A", "A", "A", "A", "A", "A", "A", "B", "B", "B", "B", "C", "D", "E", "F", "G", "H", "J"],
            "S3": [None, "A", "A", "A", "B", "B", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "P"],
            "S4": [None, "A", "A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "P", "Q", "R", "R"],
            "GI": [None, "A", "A", "A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "P", "P", "P"],
            "GII": [None, "A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "P", "Q", "R", "R", "R"],
            "GIII": [None, "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "P", "Q", "R", "R", "S", "S"],
        },
        "samples": {
            "A": 2, "B": 3, "C": 5, "D": 8, "E": 13, "F": 20, "G": 32, "H": 50,
            "J": 80, "K": 125, "L": 200, "M": 315, "N": 500, "P": 800, "Q": 1250, "R": 2000, "S": 3150,
        },
    }


@router.post("/change-password")
def change_password(
    old_password: str = Query(...),
    new_password: str = Query(..., min_length=6),
    db: Session = Depends(get_db),
    token: dict = Depends(get_current_user),
):
    from app.auth.security import hash_password, verify_password
    user = db.query(User).filter(User.id == int(token.get("sub", 0))).first()
    if not user:
        raise HTTPException(status_code=404, detail="Khong tim thay nguoi dung")
    if not verify_password(old_password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Mat khau cu khong dung")
    user.hashed_password = hash_password(new_password)
    db.commit()
    return {"message": "Doi mat khau thanh cong"}


@router.post("/import/materials")
def import_materials(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    token: dict = Depends(get_current_user),
):
    if token.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Chi admin moi co quyen")
    from io import BytesIO
    from openpyxl import load_workbook
    from app.models.iqc import Material
    contents = file.file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code, name, spec, unit, supplier_id = row[0], row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else "", row[3] if len(row) > 3 else "pcs", row[4] if len(row) > 4 else 1
        if not code:
            continue
        existing = db.query(Material).filter(Material.code == str(code).strip()).first()
        if existing:
            continue
        mat = Material(code=str(code).strip(), name=str(name).strip(), specification=str(spec).strip() if spec else None, unit=str(unit).strip() if unit else "pcs", supplier_id=int(supplier_id) if supplier_id else 1)
        db.add(mat)
        count += 1
    db.commit()
    return {"message": f"Da import {count} nguyen lieu"}


@router.post("/import/products")
def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    token: dict = Depends(get_current_user),
):
    if token.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Chi admin moi co quyen")
    from io import BytesIO
    from openpyxl import load_workbook
    from app.models.oqc import Product
    contents = file.file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code, name, spec, unit = row[0], row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else "", row[3] if len(row) > 3 else "pcs"
        if not code:
            continue
        existing = db.query(Product).filter(Product.code == str(code).strip()).first()
        if existing:
            continue
        prod = Product(code=str(code).strip(), name=str(name).strip(), specification=str(spec).strip() if spec else None, unit=str(unit).strip() if unit else "pcs")
        db.add(prod)
        count += 1
    db.commit()
    return {"message": f"Da import {count} san pham"}
