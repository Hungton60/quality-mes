from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy.orm import Session

from app.auth.security import get_current_user
from app.core.database import get_db
from app.models.iqc import IQCInspection
from app.models.oqc import OQCInspection
from app.models.ipqc import IPQCInspection

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/inspections")
def export_inspections(
    module: str = Query(..., pattern="^(iqc|oqc|ipqc)$"),
    db: Session = Depends(get_db),
    _token: dict = Depends(get_current_user),
):
    model_map = {"iqc": IQCInspection, "oqc": OQCInspection, "ipqc": IPQCInspection}
    Model = model_map[module]
    inspections = db.query(Model).order_by(Model.created_at.desc()).limit(500).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{module.upper()} Inspections"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["So phieu", "Trang thai", "Ngay kiem", "Nguoi kiem", "Ghi chu"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_labels = {"pending": "Cho kiem", "pass": "Dat", "fail": "Khong dat"}
    for row_idx, insp in enumerate(inspections, 2):
        inspector_name = insp.inspector.full_name if insp.inspector else str(insp.inspector_id)
        values = [
            insp.inspection_no,
            status_labels.get(insp.status, insp.status),
            str(insp.inspection_date)[:10],
            inspector_name,
            insp.notes or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 40

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{module}_inspections_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
